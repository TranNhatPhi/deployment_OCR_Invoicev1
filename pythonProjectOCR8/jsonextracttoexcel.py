import os

from sympy import false

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
import cv2
from paddleocr import PaddleOCR
from openpyxl import Workbook
from Preprocess import preprocess  # Import preprocess function
from openpyxl.utils import get_column_letter

# Initialize PaddleOCR
paddleOCR = PaddleOCR(lang="en",  # English language model
                      use_angle_cls=True,  # Enable angle classification for rotated text
                      show_log=False,  # Disable detailed logging for better performance
                      det_db_thresh=0.5,  # Higher sensitivity for detecting text
                      det_db_unclip_ratio=2.3,  # Adjust bounding box size for better text capture
                      drop_score=0.1,  # Filter out low-confidence text recognition results

                      )

# Đọc hình ảnh
image_path = 'image6/page_1.PNG'  # Đường dẫn tới hình ảnh bạn đã tải lên
image = cv2.imread(image_path, cv2.IMREAD_COLOR)
preprocessed_img = preprocess(image)

# Thực hiện OCR để lấy văn bản và bounding box
result = paddleOCR.ocr(image_path)

# Tạo file Excel mới
wb = Workbook()
ws = wb.active
ws.title = "OCR Results"

# Hệ số thu nhỏ để giảm khoảng cách giữa các bounding box
scaling_factor_width = 30  # Điều chỉnh để thu hẹp khoảng cách chiều rộng
scaling_factor_height = 15   # Điều chỉnh để thu hẹp khoảng cách chiều cao

# Cấu hình chiều cao và chiều rộng của các cột để phù hợp với bbox
def set_column_row_sizes(ws, bbox):
    try:
        # Đảm bảo bbox là một danh sách và chứa danh sách các tọa độ
        if isinstance(bbox, list) and all(isinstance(coord, list) for coord in bbox):
            # Xử lý từng tọa độ trong bbox
            x_min, y_min = min([pt[0] for pt in bbox if isinstance(pt, list) and len(pt) == 2]), min([pt[1] for pt in bbox if isinstance(pt, list) and len(pt) == 2])
            x_max, y_max = max([pt[0] for pt in bbox if isinstance(pt, list) and len(pt) == 2]), max([pt[1] for pt in bbox if isinstance(pt, list) and len(pt) == 2])

            # Tính toán kích thước và vị trí dựa trên hệ số thu nhỏ
            col_letter = get_column_letter(int(x_min // scaling_factor_width + 1))
            row_num = int(y_min // scaling_factor_height + 1)

            # Đặt chiều cao cho hàng
            ws.row_dimensions[row_num].height = max(ws.row_dimensions[row_num].height, (y_max - y_min) / scaling_factor_height)

            # Đặt chiều rộng cho cột
            ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width, (x_max - x_min) / scaling_factor_width)
        else:
            print(f"Invalid bbox format: {bbox}")
    except Exception as e:
        print(f"Error processing bbox: {bbox}, error: {e}")

# Lưu trữ kết quả OCR vào Excel
for idx, line in enumerate(result):
    # Lấy từng dòng kết quả với bbox và text
    for text_data in line:
        bbox, (text, confidence) = text_data

        # Đặt văn bản vào vị trí tương ứng trong file Excel
        try:
            # Xác định vị trí của các bounding box
            x_min, y_min = min([pt[0] for pt in bbox if isinstance(pt, list) and len(pt) == 2]), min([pt[1] for pt in bbox if isinstance(pt, list) and len(pt) == 2])
            col_letter = get_column_letter(int(x_min // scaling_factor_width + 1))
            row_num = int(y_min // scaling_factor_height + 1)

            # Đặt văn bản vào vị trí tương ứng
            ws[f"{col_letter}{row_num}"] = text
            print(text)  # In ra văn bản đã trích xuất để kiểm tra
            # Gọi hàm cấu hình chiều cao và chiều rộng dựa trên bbox
            set_column_row_sizes(ws, bbox)
        except Exception as e:
            print(f"Error processing text data for {text_data}: {e}")

# Lưu file Excel
output_excel_path = 'ocr_results_with_bbox_all_texts4.xlsx'
wb.save(output_excel_path)

print(f"OCR results with bounding boxes saved to {output_excel_path}")
